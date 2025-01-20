import sys
import win32com.client as win32
import os
from openpyxl import load_workbook

"""
    use 
        pip install pywin32 openpyxl
    to install necessary libraries.

"""

def reorder_sheets(input_path, sheet_name_to_move, target_position):
    workbook = load_workbook(input_path)
    sheet_names = workbook.sheetnames

    if sheet_name_to_move not in sheet_names:
        print(f"Error: Sheet '{sheet_name_to_move}' not found in the workbook.")
        return

    sheet = workbook[sheet_name_to_move]
    workbook._sheets.remove(sheet)
    workbook._sheets.insert(target_position, sheet)
    workbook.save(input_path)
    
    
try:
    if len(sys.argv) != 2:
        print("Usage: python script.py <input_file_path>")
        sys.exit(1)

    input_path = sys.argv[1]

    if os.path.exists(input_path):
        excel = win32.Dispatch("Excel.Application")
    
        excel.Visible = False
        new_workbook = excel.Workbooks.Open(input_path)

        data_sheet = new_workbook.Sheets("bank-transactions")
        last_row = data_sheet.UsedRange.Rows.Count
        last_col = data_sheet.UsedRange.Columns.Count
        data_range = data_sheet.Range(data_sheet.Cells(1, 1), data_sheet.Cells(last_row, last_col))

        pivot_cache = new_workbook.PivotCaches().Create(SourceType=1, SourceData=data_range)

        pivot_sheet = new_workbook.Sheets.Add(After=data_sheet)
        pivot_sheet.Name = "pivot-table"

        pivot_sheet.Tab.Color = 65280
        pivot_table = pivot_cache.CreatePivotTable(
            TableDestination=pivot_sheet.Cells(1, 1),
            TableName="Transaction PivotTable"
        )

        pivot_table.PivotFields("ACCOUNT NAME").Orientation = 1  # Row field
        pivot_table.PivotFields("CURRENCY").Orientation = 2      # Column field
        pivot_table.PivotFields("TRANS. AMOUNT").Orientation = 4  # Data field
        pivot_table.PivotFields("TRANS. YEAR").Orientation = 3    # Filter field
        pivot_table.PivotFields("TRANS TYPE").Orientation = 3     # Filter field

        pivot_table.TableStyle2 = "PivotStyleMedium2"
        pivot_table.DataBodyRange.NumberFormat = "#,##0.00"

        new_workbook.Save()
        new_workbook.Close()
        excel.Quit()
        
        reorder_sheets(input_path, 'pivot-table', 6)

        print(f"Advanced pivot table created successfully in '{input_path}'.")
    else:
        print("Error: The output Excel file was not created.")

except Exception as e:
    print(f"Error: {e}", file=sys.stderr)

